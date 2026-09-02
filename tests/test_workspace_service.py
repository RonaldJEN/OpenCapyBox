"""Persistent workspace path, revision, idempotency, and lifecycle tests."""

from __future__ import annotations

import asyncio
import hashlib
import json
from contextlib import asynccontextmanager
from datetime import timedelta
from types import SimpleNamespace
from unittest.mock import AsyncMock

import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from src.api.models.auth_user import AuthUser
from src.api.models.session import Session
from src.api.models.user_run_lock import UserRunLock
from src.api.models.user_sandbox import UserSandbox
from src.api.models.workspace import (
    UserWorkspace,
    WorkspaceChangeSet,
    WorkspaceClaim,
    WorkspaceContentObject,
    WorkspaceContentReference,
    WorkspaceEntry,
    WorkspaceFileVersion,
    WorkspaceMutation,
)
from src.api.services.workspace_service import (
    SandboxFileStat,
    WorkspaceError,
    WorkspacePathPolicy,
    WorkspaceService,
    WorkspaceStageResult,
    WorkspaceStore,
    _blank_xlsx_bytes,
    ensure_workspace_profile_switch_allowed,
)
from src.api.services.workspace_mutation_coordinator import (
    WorkspaceClaimLease,
    WorkspaceClaimSpec,
    WorkspaceMutationCoordinator,
    file_scope,
)
from src.api.services.file_preview_service import office_preview_cache_keys
from src.api.services.spreadsheet_edit_validation import validate_xlsx_edit_payload
from src.api.utils.timezone import now_naive


class FakeStore:
    absolute_path = WorkspaceStore.absolute_path

    def __init__(self):
        self.files: dict[str, bytes] = {}
        self.directories: set[str] = set()
        self.external: dict[tuple[str, str], tuple[bytes, str]] = {}
        self.external_directories: set[tuple[str, str]] = set()
        self.stream_read_calls = 0
        self.snapshot_cleanups: list[str] = []
        self.move_many_calls: list[list[dict[str, str]]] = []
        self.cutover_calls: list[str] = []
        self.purge_cleanup_calls: list[str] = []
        self.preview_cache_removals: list[set[str]] = []
        self.sandbox = SimpleNamespace()
        self.workspace_root = "/home/user/workdir"

    @asynccontextmanager
    async def claim_fence(self, _leases):
        yield self

    async def advance_claim_fences(self, _previous, current):
        return tuple(current)

    async def mkdir(self, path: str):
        if path in self.directories or path in self.files:
            raise WorkspaceError(409, "NAME_CONFLICT", "exists")
        self.directories.add(path)

    async def write_bytes_atomic(
        self,
        path,
        content,
        *,
        expected_sha256=None,
        must_not_exist=False,
        temp_token=None,
    ):
        current = self.files.get(path)
        if must_not_exist and current is not None:
            raise WorkspaceError(409, "NAME_CONFLICT", "exists")
        if expected_sha256 is not None:
            current_sha = hashlib.sha256(current or b"").hexdigest() if current is not None else None
            if current_sha != expected_sha256:
                raise WorkspaceError(409, "DESTINATION_CHANGED", "changed")
        self.files[path] = bytes(content)
        return self._stat(content)

    async def stage_bytes_for_install(self, content, *, temp_token):
        path = f".opencapybox/tmp/{temp_token}.bytes-staged"
        self.files[path] = bytes(content)
        return path

    async def stage_upload_stream(self, source, *, max_bytes):
        chunks = []
        total = 0
        while True:
            chunk = await source.read(4)
            self.stream_read_calls += 1
            if not chunk:
                break
            chunks.append(chunk)
            total += len(chunk)
            if total > max_bytes:
                raise WorkspaceError(413, "FILE_TOO_LARGE", "too large")
        content = b"".join(chunks)
        path = f".opencapybox/tmp/fake-upload-{self.stream_read_calls}.tmp"
        self.files[path] = content
        return path, self._stat(content)

    async def write_upload_stream_atomic(
        self,
        path,
        source,
        *,
        max_bytes,
        before_install,
        must_not_exist=False,
    ):
        staged_path, staged = await self.stage_upload_stream(source, max_bytes=max_bytes)
        before_install(staged, staged_path)
        return await self.install_staged_file(
            staged_relative_path=staged_path,
            destination_relative_path=path,
            expected_destination_sha256=None,
            must_not_exist=must_not_exist,
        )

    async def move(self, source, destination, *, allow_system_destination=False):
        if destination in self.files or destination in self.directories:
            raise WorkspaceError(409, "NAME_CONFLICT", "exists")
        if source in self.files:
            self.files[destination] = self.files.pop(source)
            return
        if source in self.directories:
            moving_directories = [path for path in self.directories if path == source or path.startswith(source + "/")]
            moving_files = [path for path in self.files if path.startswith(source + "/")]
            for path in moving_directories:
                self.directories.remove(path)
                self.directories.add(destination + path[len(source):])
            for path in moving_files:
                self.files[destination + path[len(source):]] = self.files.pop(path)
            return
        raise WorkspaceError(404, "NOT_FOUND", "missing")

    async def delete_entries(self, paths, cleanup_paths):
        for path in [*paths, *cleanup_paths]:
            try:
                await self.remove(path)
            except WorkspaceError as exc:
                if exc.code != "NOT_FOUND":
                    raise

    async def remove(self, path):
        if path in self.files:
            self.files.pop(path)
            return
        found = path in self.directories
        self.directories = {item for item in self.directories if item != path and not item.startswith(path + "/")}
        self.files = {key: value for key, value in self.files.items() if not key.startswith(path + "/")}
        if not found:
            raise WorkspaceError(404, "NOT_FOUND", "missing")

    async def remove_content_object(self, path):
        await self.remove(path)

    async def remove_content_objects(self, paths):
        for path in paths:
            try:
                await self.remove(path)
            except WorkspaceError as exc:
                if exc.code != "NOT_FOUND":
                    raise

    async def remove_office_preview_caches(self, cache_keys):
        self.preview_cache_removals.append(set(cache_keys))

    async def stat(self, path):
        if path not in self.files:
            raise WorkspaceError(404, "NOT_FOUND", "missing")
        return self._stat(self.files[path])

    async def inspect_path(self, path, *, allow_system=False):
        from src.api.services.workspace_service import WorkspacePathState

        if path in self.files:
            return WorkspacePathState("file", hashlib.sha256(self.files[path]).hexdigest())
        if path in self.directories:
            return WorkspacePathState("directory", None)
        raise WorkspaceError(404, "NOT_FOUND", "missing")

    async def stat_external(self, root, path):
        if root == self.workspace_root:
            return await self.stat(path)
        try:
            content, revision = self.external[(root, path)]
        except KeyError as exc:
            raise WorkspaceError(404, "NOT_FOUND", "missing") from exc
        size, mtime_ns = revision.split(":")[1:]
        return SandboxFileStat(int(size), int(mtime_ns), hashlib.sha256(content).hexdigest())

    async def read_bytes(self, path, *, allow_system=False):
        return bytes(self.files[path])

    async def copy_external_atomic(
        self,
        *,
        source_root,
        source_relative_path,
        expected_source_revision,
        destination_relative_path,
        expected_destination_sha256=None,
        must_not_exist=False,
        temp_token=None,
        allow_system_destination=False,
    ):
        if source_root == self.workspace_root:
            content = self.files[source_relative_path]
            revision = self._stat(content).source_revision
        else:
            content, revision = self.external[(source_root, source_relative_path)]
        if revision != expected_source_revision:
            raise WorkspaceError(412, "SOURCE_REVISION_CONFLICT", "changed")
        return await self.write_bytes_atomic(
            destination_relative_path,
            content,
            expected_sha256=expected_destination_sha256,
            must_not_exist=must_not_exist,
        )

    async def install_staged_file(
        self,
        *,
        staged_relative_path,
        destination_relative_path,
        expected_destination_sha256,
        must_not_exist,
    ):
        content = self.files.pop(staged_relative_path)
        return await self.write_bytes_atomic(
            destination_relative_path,
            content,
            expected_sha256=expected_destination_sha256,
            must_not_exist=must_not_exist,
        )

    async def copy_to_external_atomic(
        self,
        *,
        source_relative_path,
        expected_source_sha256,
        destination_root,
        destination_relative_path,
    ):
        content = self.files[source_relative_path]
        assert hashlib.sha256(content).hexdigest() == expected_source_sha256
        self.external[(destination_root, destination_relative_path)] = (
            content,
            f"v1:{len(content)}:1",
        )
        return self._stat(content)

    async def ensure_external_directory(
        self,
        *,
        destination_root,
        destination_relative_path,
        must_not_exist=False,
    ):
        key = (destination_root, destination_relative_path)
        if must_not_exist and key in self.external_directories:
            raise WorkspaceError(409, "NAME_CONFLICT", "exists")
        self.external_directories.add(key)

    async def inspect_external_directory_manifest(
        self,
        *,
        destination_root,
        directory_relative_path,
    ):
        prefix = directory_relative_path + "/"
        manifest = []
        for root, path in sorted(self.external_directories):
            if root == destination_root and path.startswith(prefix):
                manifest.append({
                    "path": path[len(prefix):],
                    "kind": "directory",
                })
        for (root, path), (content, _revision) in sorted(self.external.items()):
            if root == destination_root and path.startswith(prefix):
                manifest.append({
                    "path": path[len(prefix):],
                    "kind": "file",
                    "sha256": hashlib.sha256(content).hexdigest(),
                    "size_bytes": len(content),
                })
        return sorted(manifest, key=lambda item: item["path"])

    async def install_external_directory_atomic(
        self,
        *,
        destination_root,
        staged_relative_path,
        destination_relative_path,
        expected_manifest,
    ):
        observed = await self.inspect_external_directory_manifest(
            destination_root=destination_root,
            directory_relative_path=staged_relative_path,
        )
        assert observed == sorted(expected_manifest, key=lambda item: item["path"])
        destination_key = (destination_root, destination_relative_path)
        if destination_key in self.external_directories:
            raise WorkspaceError(409, "NAME_CONFLICT", "exists")
        staged_prefix = staged_relative_path + "/"
        destination_prefix = destination_relative_path + "/"
        moving_directories = [
            (root, path)
            for root, path in self.external_directories
            if root == destination_root and (path == staged_relative_path or path.startswith(staged_prefix))
        ]
        moving_files = [
            (root, path)
            for root, path in self.external
            if root == destination_root and path.startswith(staged_prefix)
        ]
        if not moving_directories:
            raise WorkspaceError(404, "NOT_FOUND", "missing")
        for key in moving_directories:
            self.external_directories.remove(key)
            suffix = key[1][len(staged_relative_path):]
            self.external_directories.add((destination_root, destination_relative_path + suffix))
        for key in moving_files:
            content = self.external.pop(key)
            suffix = key[1][len(staged_prefix):]
            self.external[(destination_root, destination_prefix + suffix)] = content

    async def cleanup_external_incoming_directory(
        self,
        *,
        destination_root,
        incoming_relative_path,
    ):
        key = (destination_root, incoming_relative_path)
        found = key in self.external_directories
        prefix = incoming_relative_path + "/"
        self.external_directories = {
            item
            for item in self.external_directories
            if not (item[0] == destination_root and (item[1] == incoming_relative_path or item[1].startswith(prefix)))
        }
        self.external = {
            item: value
            for item, value in self.external.items()
            if not (item[0] == destination_root and item[1].startswith(prefix))
        }
        return found

    async def snapshot_for_read(self, *, source_relative_path, destination_relative_path):
        content = self.files[source_relative_path]
        self.files[destination_relative_path] = content
        return SandboxFileStat(len(content), 1, None)

    async def copy_version_snapshot(
        self,
        *,
        source_relative_path,
        destination_relative_path,
        expected_sha256,
        expected_size_bytes,
    ):
        content = bytes(self.files[source_relative_path])
        stat = self._stat(content)
        assert stat.sha256 == expected_sha256
        assert stat.size_bytes == expected_size_bytes
        self.files[destination_relative_path] = content
        return stat

    async def ensure_content_object(
        self,
        *,
        source_relative_path,
        destination_relative_path,
        expected_sha256,
        expected_size_bytes,
    ):
        content = bytes(self.files[source_relative_path])
        stat = self._stat(content)
        assert stat.sha256 == expected_sha256
        assert stat.size_bytes == expected_size_bytes
        existing = self.files.get(destination_relative_path)
        if existing is not None:
            assert existing == content
        else:
            self.files[destination_relative_path] = content
        return stat

    async def restore_version_atomic(
        self,
        *,
        version_relative_path,
        destination_relative_path,
        expected_version_sha256,
        expected_version_size_bytes,
        expected_destination_sha256,
        temp_token,
    ):
        content = bytes(self.files[version_relative_path])
        assert hashlib.sha256(content).hexdigest() == expected_version_sha256
        assert len(content) == expected_version_size_bytes
        return await self.write_bytes_atomic(
            destination_relative_path,
            content,
            expected_sha256=expected_destination_sha256,
            temp_token=temp_token,
        )

    async def cleanup_read_snapshot_links(self, path):
        self.snapshot_cleanups.append(path)
        return 0

    @staticmethod
    def _stat(content: bytes):
        return SandboxFileStat(len(content), 1, hashlib.sha256(content).hexdigest())


class FakeSandboxService:
    def __init__(self, mount_path="/home/user"):
        self.mount_path = mount_path

    def get_mount_path(self, _user_id=None):
        return self.mount_path


class SandboxLookupDB:
    def __init__(self, sandbox_id: str | None):
        self.row = SimpleNamespace(sandbox_id=sandbox_id) if sandbox_id else None

    def query(self, _model):
        row = self.row

        class Query:
            def filter(self, *_conditions):
                return self

            def first(self):
                return row

        return Query()

    def commit(self):
        return None


class ServiceUnderTest(WorkspaceService):
    def __init__(self, db, workspace, store):
        self.db = db
        self.workspace = workspace
        self.store = store
        self.sandbox_service = FakeSandboxService()
        self.settings = SimpleNamespace(
            workspace_quota_bytes=1024 * 1024,
            workspace_history_quota_bytes=1024 * 1024,
            workspace_preview_cache_bytes=1024 * 1024,
            workspace_max_file_bytes=1024 * 1024,
            workspace_max_entries=100,
            workspace_mutation_lease_seconds=120,
            workspace_version_retention_count=20,
            workspace_version_retention_days=30,
            workspace_draft_base_retention_days=1,
            workspace_draft_revision_retention_count=5,
            workspace_history_gc_interval_seconds=300,
            workspace_history_gc_batch_size=100,
        )

    async def _prepare(
        self,
        user_id: str,
        *,
        for_update: bool,
        reconcile_prepared: bool = True,
        require_filesystem: bool = True,
    ):
        workspace = self.db.query(UserWorkspace).filter(UserWorkspace.user_id == user_id).one()
        return workspace, self.store if require_filesystem else None


class CrashAfterFilesystemService(ServiceUnderTest):
    def _record_mutation(self, **kwargs):
        if kwargs.get("prepared_mutation") is not None:
            self.db.rollback()
            raise RuntimeError("simulated database outage after filesystem mutation")
        return super()._record_mutation(**kwargs)


class ProfileFenceService(WorkspaceService):
    def __init__(self, db, runtime):
        self.db = db
        self.runtime = runtime
        self.sandbox_service = FakeSandboxService()
        self.settings = SimpleNamespace(
            sandbox_persistent_storage_enabled=True,
            workspace_quota_bytes=1024 * 1024,
        )
        self.sandbox_requested = False

    def _runtime(self, user_id: str):
        return self.runtime

    async def _sandbox_for_user(self, user_id: str):
        self.sandbox_requested = True
        raise AssertionError("profile fence must run before sandbox lifecycle")


@pytest.fixture
def workspace_db():
    engine = create_engine("sqlite:///:memory:")
    AuthUser.__table__.create(engine)
    UserWorkspace.__table__.create(engine)
    WorkspaceEntry.__table__.create(engine)
    WorkspaceContentObject.__table__.create(engine)
    WorkspaceContentReference.__table__.create(engine)
    WorkspaceFileVersion.__table__.create(engine)
    WorkspaceMutation.__table__.create(engine)
    WorkspaceClaim.__table__.create(engine)
    WorkspaceChangeSet.__table__.create(engine)
    Session.__table__.create(engine)
    UserSandbox.__table__.create(engine)
    UserRunLock.__table__.create(engine)
    db = sessionmaker(bind=engine, expire_on_commit=False)()
    db.add(AuthUser(
        user_id="user-1",
        username="workspace-user-1",
        auth_type="simple",
        password_hash="test",
        enabled=True,
    ))
    workspace = UserWorkspace(
        user_id="user-1",
        root_path="/home/user/workdir",
        active_profile_id="profile-1",
        active_profile_version=1,
        quota_bytes=1024 * 1024,
    )
    db.add(workspace)
    db.commit()
    try:
        yield db, workspace
    finally:
        db.close()
        engine.dispose()


def expire_prepared_mutation_ownership(db) -> None:
    expired_at = now_naive() - timedelta(seconds=1)
    db.query(WorkspaceMutation).filter(
        WorkspaceMutation.state == "prepared",
    ).update({"lease_expires_at": expired_at}, synchronize_session=False)
    db.query(WorkspaceClaim).filter(
        WorkspaceClaim.state == "active",
    ).update({"lease_expires_at": expired_at}, synchronize_session=False)
    db.commit()


@pytest.mark.parametrize(
    "path",
    [
        "/absolute",
        "../escape",
        "a/../../escape",
        "a\\b",
        "a//b",
        "a/./b",
        ".opencapybox/trash/x",
        ".opencapybox-preview/cache.pdf",
    ],
)
def test_workspace_path_policy_rejects_boundary_escape(path):
    with pytest.raises(WorkspaceError):
        WorkspacePathPolicy.normalize_relative_path(path)






@pytest.mark.asyncio
async def test_content_object_publish_is_no_clobber_and_validates_concurrent_winner():
    digest = hashlib.sha256(b"same").hexdigest()
    command_run = AsyncMock(return_value=SimpleNamespace(
        exit_code=0,
        stdout=json.dumps({"ok": True, "size": 4, "mtime_ns": 1, "sha256": digest}),
    ))
    sandbox = SimpleNamespace(commands=SimpleNamespace(run=command_run))
    store = WorkspaceStore(sandbox, "/home/user/workdir")

    result = await store.ensure_content_object(
        source_relative_path="same.md",
        destination_relative_path=f".opencapybox/objects/sha256/{digest[:2]}/{digest}/content",
        expected_sha256=digest,
        expected_size_bytes=4,
    )

    assert result.sha256 == digest
    command = command_run.await_args.args[0]
    body = command.split("<<'PY'\n", 1)[1].rsplit("\nPY", 1)[0]
    compile(body, "<ensure_content_object>", "exec")
    assert "no_clobber = True" in body
    assert "os.link(" in body
    assert "except FileExistsError" in body
    assert "existing_digest.hexdigest() != source_sha" in body
    assert "os.chmod(" in body


@pytest.mark.asyncio
async def test_restore_version_copies_object_before_atomic_active_install():
    digest = hashlib.sha256(b"same").hexdigest()
    command_run = AsyncMock(return_value=SimpleNamespace(
        exit_code=0,
        stdout=json.dumps({"ok": True, "size": 4, "mtime_ns": 1, "sha256": digest}),
    ))
    sandbox = SimpleNamespace(commands=SimpleNamespace(run=command_run))
    store = WorkspaceStore(sandbox, "/home/user/workdir")

    restored = await store.restore_version_atomic(
        version_relative_path=f".opencapybox/objects/sha256/{digest[:2]}/{digest}/content",
        destination_relative_path="report.md",
        expected_version_sha256=digest,
        expected_version_size_bytes=4,
        expected_destination_sha256="old-sha",
        temp_token="restore-1",
    )

    assert restored.sha256 == digest
    copy_body = command_run.await_args_list[0].args[0].split("<<'PY'\n", 1)[1].rsplit("\nPY", 1)[0]
    install_body = command_run.await_args_list[1].args[0].split("<<'PY'\n", 1)[1].rsplit("\nPY", 1)[0]
    compile(copy_body, "<restore_version_copy>", "exec")
    compile(install_body, "<restore_version_install>", "exec")
    assert "no_clobber = False" in copy_body
    assert "os.write(out_fd" in copy_body
    assert "os.replace(temp_name, dest_name" in copy_body
    assert "os.replace(temp_name, dest_name" in install_body




@pytest.mark.asyncio
async def test_current_and_fixed_version_reads_use_immutable_objects_without_request_hash(workspace_db):
    db, workspace = workspace_db
    store = FakeStore()
    service = ServiceUnderTest(db, workspace, store)

    created = await service.upload_file("user-1", None, "报告.md", b"first")
    entry_id = created.entry.entry_id
    first_version = created.entry.current_version_id
    opened = await service.open_content("user-1", entry_id)
    await service.write_content(
        "user-1",
        entry_id,
        b"newer content",
        created.entry.revision,
    )
    latest = await service.open_content("user-1", entry_id)
    store.inspect_path = AsyncMock(wraps=store.inspect_path)
    selected = await service.open_version_content("user-1", first_version)
    assert opened.entry.revision == 1
    assert latest.entry.revision == 2
    assert opened.sandbox_path == selected.sandbox_path != latest.sandbox_path
    assert selected.version.version_id == first_version
    assert selected.version.size_bytes == 5
    for result, expected in ((opened, b"first"), (selected, b"first"), (latest, b"newer content")):
        path = result.sandbox_path.removeprefix(store.workspace_root + "/")
        assert store.files[path] == expected
    store.inspect_path.assert_not_awaited()
    assert store.snapshot_cleanups == []
    assert not any(path.startswith(".opencapybox/read/") for path in store.files)


@pytest.mark.asyncio
async def test_workspace_child_reference_preserves_entry_version_identity(workspace_db):
    from src.api.schemas.chat import AssistantFileReference
    from src.api.services.agent_service import AgentService

    db, workspace = workspace_db
    store = FakeStore()
    service = ServiceUnderTest(db, workspace, store)

    created = await service.upload_file("user-1", None, "report.md", b"child")
    entry_id, version_id = created.entry.entry_id, created.entry.current_version_id
    reference = AssistantFileReference(
        ref_id=f"workspace:{entry_id}:{version_id}", source="workspace",
        entry_id=entry_id, version_id=version_id, name="report.md", path="report.md",
        revision="1", size=5, operation="CREATED", tool_call_id="child-tool",
    ).model_dump()
    parent = object.__new__(AgentService)
    parent.user_id = "user-1"
    parent.session_id = "parent-session"
    parent.history_service = SimpleNamespace(db=db)
    parent.sandbox = store.sandbox
    materialized = await parent._materialize_assistant_file_reference(reference, run_id="parent-round")
    assert materialized["source"] == "workspace"
    assert materialized["version_id"] == version_id
    assert materialized["entry_id"] == entry_id
    assert materialized["toolCallId"] == "child-tool"
    db.flush()
    assert db.query(WorkspaceContentReference).filter_by(
        reference_kind="round_attachment",
        reference_key=f"parent-session:parent-round:assistant:{entry_id}:{version_id}",
        version_id=version_id,
    ).count() == 1
    with pytest.raises(WorkspaceError) as mismatch:
        await parent._materialize_assistant_file_reference(
            {**reference, "entry_id": "different-entry"}, run_id="parent-round",
        )
    assert mismatch.value.code == "VERSION_NOT_FOUND"
    with pytest.raises(WorkspaceError):
        await service.open_version_content("user-1", "missing-version")












@pytest.mark.asyncio
async def test_json_script_rejects_missing_exit_code_as_invalid_response():
    sandbox = SimpleNamespace(
        commands=SimpleNamespace(
            run=AsyncMock(return_value=SimpleNamespace(
                exit_code=None,
                stdout='{"ok": true}',
            ))
        )
    )
    store = WorkspaceStore(sandbox, "/home/user/workdir")

    with pytest.raises(WorkspaceError) as invalid:
        await store._run_json_script("print('{}')")

    assert invalid.value.code == "SANDBOX_RESPONSE_INVALID"


@pytest.mark.asyncio
async def test_store_claim_fence_is_held_and_validated_before_final_side_effect():
    runner = AsyncMock(return_value=SimpleNamespace(
        exit_code=0,
        stdout='{"ok": true}',
    ))
    store = WorkspaceStore(
        SimpleNamespace(commands=SimpleNamespace(run=runner)),
        "/home/user/workdir",
    )
    lease = WorkspaceClaimLease(
        claim_id="claim-1",
        user_id="user-1",
        scope_kind="file",
        scope_key="file:entry-1",
        owner_token="owner-1",
        generation=3,
        lease_expires_at=now_naive(),
    )

    async with store.claim_fence((lease,)):
        await store._run_json_script(
            "import json, os\nos.replace('source.tmp', 'target.md')\n"
            "print(json.dumps({'ok': True}))"
        )

    command = runner.await_args.args[0]
    body = command.split("<<'PY'\n", 1)[1].rsplit("\nPY", 1)[0]
    compile(body, "<workspace-fenced-side-effect>", "exec")
    assert body.index("fcntl.flock(lock_fd, fcntl.LOCK_EX)") < body.index(
        "os.replace('source.tmp', 'target.md')"
    )
    assert body.index("current_generation > expected_generation") < body.index(
        "os.replace('source.tmp', 'target.md')"
    )




def test_blank_xlsx_contains_a_valid_sheet1():
    from io import BytesIO

    from openpyxl import load_workbook

    content = _blank_xlsx_bytes()
    validate_xlsx_edit_payload(content)
    workbook = load_workbook(BytesIO(content), read_only=True)
    try:
        assert workbook.active.calculate_dimension(force=True) == "A1:A1"
    finally:
        workbook.close()


def _disconnected_xlsx_bytes() -> bytes:
    from io import BytesIO
    from zipfile import ZipFile

    output = BytesIO()
    with ZipFile(output, "w") as archive:
        archive.writestr("[Content_Types].xml", "not-content-types")
        archive.writestr("xl/workbook.xml", "not-a-workbook")
    return output.getvalue()


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("name", "original", "invalid", "error_code"),
    [
        ("report.csv", b"name,value\na,1\n", b"name,value\nhello,\x00world\n", "INVALID_CSV"),
        ("report.xlsx", _blank_xlsx_bytes(), _disconnected_xlsx_bytes(), "INVALID_XLSX"),
    ],
)
async def test_workspace_edit_rejects_invalid_spreadsheet_without_mutation(
    workspace_db,
    name,
    original,
    invalid,
    error_code,
):
    db, _workspace = workspace_db
    store = FakeStore()
    service = ServiceUnderTest(db, _workspace, store)
    created = await service.upload_file("user-1", None, name, original)
    mutation_count = db.query(WorkspaceMutation).count()
    version_count = db.query(WorkspaceFileVersion).count()

    with pytest.raises(WorkspaceError) as invalid_edit:
        await service.write_content(
            "user-1",
            created.entry.entry_id,
            invalid,
            expected_revision=created.entry.revision,
        )

    assert invalid_edit.value.code == error_code
    assert store.files[name] == original
    current = await service.get_entry("user-1", created.entry.entry_id)
    assert current.revision == created.entry.revision
    assert current.current_version_id == created.entry.current_version_id
    assert db.query(WorkspaceMutation).count() == mutation_count
    assert db.query(WorkspaceFileVersion).count() == version_count


@pytest.mark.asyncio
async def test_workspace_auto_merge_validates_before_head_synchronization(workspace_db):
    db, workspace = workspace_db
    store = FakeStore()
    service = ServiceUnderTest(db, workspace, store)
    created = await service.upload_file(
        "user-1",
        None,
        "report.xlsx",
        _blank_xlsx_bytes(),
    )
    service._synchronize_physical_head = AsyncMock()

    with pytest.raises(WorkspaceError) as invalid_edit:
        await service.write_content_auto_merge(
            "user-1",
            created.entry.entry_id,
            _disconnected_xlsx_bytes(),
            created.entry.revision,
            base_version_id=created.entry.current_version_id,
        )

    assert invalid_edit.value.code == "INVALID_XLSX"
    service._synchronize_physical_head.assert_not_awaited()


def test_xlsx_validation_rejects_uncompressed_budget(monkeypatch):
    from src.api.services import spreadsheet_edit_validation

    monkeypatch.setattr(
        spreadsheet_edit_validation,
        "MAX_XLSX_UNCOMPRESSED_BYTES",
        1,
    )
    with pytest.raises(ValueError, match="解压后内容过大"):
        spreadsheet_edit_validation.validate_xlsx_edit_payload(_blank_xlsx_bytes())




@pytest.mark.asyncio
async def test_file_versions_are_immutable_and_restore_creates_a_new_head(workspace_db):
    db, workspace = workspace_db
    store = FakeStore()
    service = ServiceUnderTest(db, workspace, store)

    created = await service.upload_file("user-1", None, "history.md", b"version one")
    first_version_id = created.entry.current_version_id
    updated = await service.write_content(
        "user-1",
        created.entry.entry_id,
        b"version two",
        expected_revision=created.entry.revision,
    )
    entry_id = updated.entry.entry_id
    updated_version_id = updated.entry.current_version_id
    versions = await service.list_versions("user-1", entry_id)
    assert [item.sequence for item in versions] == [2, 1]
    assert first_version_id == versions[1].version_id

    restored = await service.restore_version(
        "user-1",
        entry_id,
        first_version_id,
        expected_revision=updated.entry.revision,
        idempotency_key="restore-history-v1",
    )

    assert restored.entry.revision == 3
    assert restored.entry.current_version_id not in {first_version_id, updated_version_id}
    assert store.files["history.md"] == b"version one"
    restored_head = db.get(WorkspaceFileVersion, restored.entry.current_version_id)
    assert restored_head.sequence == 3
    assert restored_head.restored_from_version_id == first_version_id
    assert store.files[versions[1].content_path] == b"version one"
    assert store.files[versions[0].content_path] == b"version two"


@pytest.mark.asyncio
async def test_existing_file_save_requires_base_and_stale_base_three_way_merges(workspace_db):
    db, workspace = workspace_db
    store = FakeStore()
    service = ServiceUnderTest(db, workspace, store)
    base = b"Human: old\nAgent: old\n"
    server = b"Human: old\nAgent: generated\n"
    draft = b"Human: edited\nAgent: old\n"
    created = await service.upload_file("user-1", None, "shared.md", base)
    base_version_id = created.entry.current_version_id
    base_revision = int(created.entry.revision)
    await service.write_content(
        "user-1",
        created.entry.entry_id,
        server,
        expected_revision=base_revision,
    )

    with pytest.raises(WorkspaceError) as missing_base:
        await service.write_content_auto_merge(
            "user-1",
            created.entry.entry_id,
            draft,
            base_revision,
            base_version_id=None,
        )
    assert missing_base.value.code == "BASE_VERSION_REQUIRED"

    with pytest.raises(WorkspaceError) as unknown_base:
        await service.write_content_auto_merge(
            "user-1",
            created.entry.entry_id,
            draft,
            base_revision,
            base_version_id="missing-base-version",
        )
    assert unknown_base.value.code == "BASE_VERSION_INVALID"

    other = await service.upload_file("user-1", None, "other.md", b"other")
    with pytest.raises(WorkspaceError) as cross_entry_base:
        await service.write_content_auto_merge(
            "user-1",
            created.entry.entry_id,
            draft,
            base_revision,
            base_version_id=other.entry.current_version_id,
        )
    assert cross_entry_base.value.code == "BASE_VERSION_INVALID"
    assert store.files["shared.md"] == server

    merged = await service.write_content_auto_merge(
        "user-1",
        created.entry.entry_id,
        draft,
        base_revision,
        base_version_id=base_version_id,
    )

    assert merged.auto_merged is True
    assert store.files["shared.md"] == b"Human: edited\nAgent: generated\n"


@pytest.mark.asyncio
async def test_stale_web_save_only_degrades_when_known_base_was_pruned(workspace_db):
    db, workspace = workspace_db
    store = FakeStore()
    service = ServiceUnderTest(db, workspace, store)
    created = await service.upload_file("user-1", None, "pruned.md", b"base")
    base_version_id = created.entry.current_version_id
    base_revision = int(created.entry.revision)
    await service.write_content(
        "user-1",
        created.entry.entry_id,
        b"server",
        expected_revision=base_revision,
    )
    base_version = db.get(WorkspaceFileVersion, base_version_id)
    base_version.state = "pruned"
    base_version.blob_id = None
    base_version.content_path = None
    base_version.pruned_at = now_naive()
    db.commit()

    saved = await service.write_content_auto_merge(
        "user-1",
        created.entry.entry_id,
        b"human draft",
        base_revision,
        base_version_id=base_version_id,
    )

    assert saved.auto_merged is True
    assert store.files["pruned.md"] == b"human draft"


@pytest.mark.asyncio
async def test_content_objects_deduplicate_same_user_same_sha(workspace_db):
    db, workspace = workspace_db
    store = FakeStore()
    service = ServiceUnderTest(db, workspace, store)

    first = await service.upload_file("user-1", None, "one.md", b"same")
    second = await service.upload_file("user-1", None, "two.md", b"same")

    first_version = db.get(WorkspaceFileVersion, first.entry.current_version_id)
    second_version = db.get(WorkspaceFileVersion, second.entry.current_version_id)
    assert first_version.blob_id == second_version.blob_id
    assert first_version.content_path == second_version.content_path
    assert db.query(WorkspaceContentObject).count() == 1
    assert store.files[first_version.content_path] == b"same"
    db.refresh(workspace)
    assert workspace.history_used_bytes == len(b"same")


@pytest.mark.asyncio
async def test_identical_save_records_no_change_without_new_revision_or_blob(workspace_db):
    db, workspace = workspace_db
    store = FakeStore()
    service = ServiceUnderTest(db, workspace, store)
    created = await service.upload_file("user-1", None, "same.md", b"same")
    original_prepare = service._prepare

    async def prepare_with_detached_workspace(*args, **kwargs):
        prepared_workspace, prepared_store = await original_prepare(*args, **kwargs)
        db.expunge(prepared_workspace)
        return prepared_workspace, prepared_store

    service._prepare = prepare_with_detached_workspace

    result = await service.write_content(
        "user-1",
        created.entry.entry_id,
        b"same",
        expected_revision=created.entry.revision,
        idempotency_key="same-content",
    )

    assert result.status == "NO_CHANGE"
    assert result.entry.revision == created.entry.revision
    assert db.query(WorkspaceFileVersion).count() == 1
    assert db.query(WorkspaceContentObject).count() == 1


@pytest.mark.asyncio
async def test_web_autosave_revision_is_hidden_until_checkpoint(workspace_db):
    db, workspace = workspace_db
    store = FakeStore()
    service = ServiceUnderTest(db, workspace, store)
    created = await service.upload_file("user-1", None, "draft.md", b"v1")
    second = await service.write_content(
        "user-1",
        created.entry.entry_id,
        b"v2",
        expected_revision=created.entry.revision,
    )
    third = await service.write_content(
        "user-1",
        created.entry.entry_id,
        b"v3",
        expected_revision=second.entry.revision,
    )

    visible_before = await service.list_versions("user-1", created.entry.entry_id)
    assert [item.sequence for item in visible_before] == [3, 1]
    promoted = await service.checkpoint_entry(
        "user-1",
        created.entry.entry_id,
        expected_revision=third.entry.revision,
        version_id=third.entry.current_version_id,
        checkpoint_kind="web_idle",
    )
    assert promoted.checkpoint_kind == "web_idle"
    assert promoted.retained_until is None
    visible_after = await service.list_versions("user-1", created.entry.entry_id)
    assert [item.sequence for item in visible_after] == [3, 1]


@pytest.mark.asyncio
async def test_history_gc_keeps_current_and_prunes_old_unreferenced_checkpoints(workspace_db):
    db, workspace = workspace_db
    store = FakeStore()
    service = ServiceUnderTest(db, workspace, store)
    service.settings.workspace_version_retention_count = 1
    service.settings.workspace_version_retention_days = 1
    created = await service.upload_file("user-1", None, "gc.md", b"v1")
    current = created
    for index in range(2, 5):
        current = await service.write_content(
            "user-1",
            created.entry.entry_id,
            f"v{index}".encode(),
            expected_revision=current.entry.revision,
        )
        await service.checkpoint_entry(
            "user-1",
            created.entry.entry_id,
            expected_revision=current.entry.revision,
            version_id=current.entry.current_version_id,
            checkpoint_kind="web_periodic",
        )
    old = now_naive() - timedelta(days=60)
    db.query(WorkspaceFileVersion).update(
        {
            WorkspaceFileVersion.created_at: old,
            WorkspaceFileVersion.retained_until: None,
        },
        synchronize_session=False,
    )
    db.commit()

    result = await service.run_history_gc("user-1", at=now_naive(), limit=20)

    assert result.versions_pruned == 3
    assert result.objects_pruned == 3
    assert result.bytes_reclaimed == 6
    current_version = db.get(WorkspaceFileVersion, current.entry.current_version_id)
    assert current_version.state == "materialized"
    assert db.query(WorkspaceContentObject).filter_by(state="materialized").count() == 1
    db.refresh(workspace)
    assert workspace.history_used_bytes == len(b"v4")


@pytest.mark.asyncio
async def test_agent_cannot_delete_entry_after_same_round_publish(workspace_db):
    db, workspace = workspace_db
    store = FakeStore()
    service = ServiceUnderTest(db, workspace, store)
    created = await service.upload_file("user-1", None, "stable.md", b"base")
    proposal = b"updated"
    store.external[("/home/user", "sessions/session-1/candidate.md")] = (
        proposal,
        f"v1:{len(proposal)}:41",
    )
    context = {"session_id": "session-1", "round_id": "round-1"}
    with pytest.raises(WorkspaceError) as missing_base:
        await service.publish_sandbox_file(
            "user-1",
            source_path="/home/user/sessions/session-1/candidate.md",
            destination_parent_id=None,
            destination_name="stable.md",
            conflict_policy="overwrite",
            expected_destination_revision=created.entry.revision,
            actor="chat",
            context=context,
            idempotency_key="missing-base-publish",
        )
    assert missing_base.value.status_code == 428
    assert missing_base.value.code == "BASE_VERSION_REQUIRED"
    assert db.query(WorkspaceChangeSet).count() == 0

    published = await service.publish_sandbox_file(
        "user-1",
        source_path="/home/user/sessions/session-1/candidate.md",
        destination_parent_id=None,
        destination_name="stable.md",
        conflict_policy="overwrite",
        expected_destination_revision=created.entry.revision,
        base_version_id=created.entry.current_version_id,
        actor="chat",
        context=context,
        idempotency_key="stable-publish",
    )
    assert published.status == "APPLIED"
    current = await service.get_entry("user-1", created.entry.entry_id)

    with pytest.raises(WorkspaceError) as exc_info:
        await service.delete_entry(
            "user-1",
            current.entry_id,
            expected_revision=current.revision,
            actor="chat",
            context=context,
            idempotency_key="forbidden-trash-workaround",
        )

    assert exc_info.value.code == "STABLE_ENTRY_ID_REQUIRED"
    assert store.files["stable.md"] == proposal


@pytest.mark.asyncio
async def test_agent_cannot_recreate_same_path_after_same_round_trash(workspace_db):
    db, workspace = workspace_db
    store = FakeStore()
    service = ServiceUnderTest(db, workspace, store)
    created = await service.upload_file("user-1", None, "stable.md", b"base")
    context = {"session_id": "session-1", "round_id": "round-2"}
    await service.delete_entry(
        "user-1",
        created.entry.entry_id,
        expected_revision=created.entry.revision,
        actor="chat",
        context=context,
        idempotency_key="explicit-trash",
    )
    replacement = b"replacement"
    store.external[("/home/user", "sessions/session-1/replacement.md")] = (
        replacement,
        f"v1:{len(replacement)}:43",
    )

    with pytest.raises(WorkspaceError) as exc_info:
        await service.publish_sandbox_file(
            "user-1",
            source_path="/home/user/sessions/session-1/replacement.md",
            destination_parent_id=None,
            destination_name="stable.md",
            actor="chat",
            context=context,
            idempotency_key="forbidden-replacement",
        )

    assert exc_info.value.code == "STABLE_ENTRY_REPLACEMENT_FORBIDDEN"
    assert "stable.md" not in store.files
    assert not any(value == b"base" for value in store.files.values())


@pytest.mark.asyncio
async def test_agent_cannot_recreate_path_after_same_round_rename(workspace_db):
    db, workspace = workspace_db
    store = FakeStore()
    service = ServiceUnderTest(db, workspace, store)
    created = await service.upload_file("user-1", None, "stable.md", b"base")
    context = {"session_id": "session-1", "round_id": "round-3"}
    await service.move_entry(
        "user-1",
        created.entry.entry_id,
        parent_id=None,
        name="stable-backup.md",
        expected_revision=created.entry.revision,
        actor="chat",
        context=context,
        idempotency_key="forbidden-rename-workaround",
    )
    replacement = b"replacement"
    store.external[("/home/user", "sessions/session-1/replacement.md")] = (
        replacement,
        f"v1:{len(replacement)}:47",
    )

    with pytest.raises(WorkspaceError) as exc_info:
        await service.publish_sandbox_file(
            "user-1",
            source_path="/home/user/sessions/session-1/replacement.md",
            destination_parent_id=None,
            destination_name="stable.md",
            actor="chat",
            context=context,
            idempotency_key="forbidden-recreate-after-rename",
        )

    assert exc_info.value.code == "STABLE_ENTRY_REPLACEMENT_FORBIDDEN"
    assert store.files["stable-backup.md"] == b"base"
    assert "stable.md" not in store.files


@pytest.mark.asyncio
@pytest.mark.parametrize("extension", ["csv", "xlsx"])
@pytest.mark.parametrize("human_changed", [False, True])
async def test_agent_publish_structural_change_respects_base(workspace_db, extension, human_changed):
    from io import BytesIO
    from openpyxl import Workbook, load_workbook

    if extension == "csv":
        base = b"name,value\na,1\n"
        proposal = b"name,value\na,1\nb,2\n"
        human = b"name,value\na,99\n"
    else:
        workbook = Workbook()
        workbook.active["A1"] = "base"
        output = BytesIO()
        workbook.save(output)
        base = output.getvalue()
        workbook.create_sheet("Added")["A1"] = "new data"
        output = BytesIO()
        workbook.save(output)
        proposal = output.getvalue()
        workbook = load_workbook(BytesIO(base))
        workbook.active["A1"] = "human edit"
        output = BytesIO()
        workbook.save(output)
        human = output.getvalue()

    db, workspace = workspace_db
    store = FakeStore()
    service = ServiceUnderTest(db, workspace, store)
    name = f"report.{extension}"
    created = await service.upload_file("user-1", None, name, base)
    entry_id, base_id = created.entry.entry_id, created.entry.current_version_id
    if human_changed:
        await service.write_content("user-1", entry_id, human, expected_revision=1)
    actor = "chat" if extension == "csv" else "cron"
    source = f"sessions/session-1/{name}" if actor == "chat" else f"cron/runs/run-1/{name}"
    store.external[("/home/user", source)] = (proposal, f"v1:{len(proposal)}:99")

    result = await service.publish_sandbox_file(
        "user-1", source_path=f"/home/user/{source}", destination_parent_id=None,
        destination_name=name, conflict_policy="overwrite", expected_destination_revision=1,
        base_version_id=base_id, actor=actor, idempotency_key="publish-structure",
    )

    expected = human if human_changed else proposal
    assert result.status == "APPLIED"
    assert store.files[name] == expected
    current = await service.get_entry("user-1", entry_id)
    head = db.get(WorkspaceFileVersion, current.current_version_id)
    assert store.files[head.content_path] == expected
    assert current.sha256 == head.sha256 == hashlib.sha256(expected).hexdigest()
    if extension == "xlsx":
        parsed = load_workbook(BytesIO(store.files[name]))
        assert ("Added" in parsed.sheetnames) is not human_changed


@pytest.mark.asyncio
async def test_xlsx_merge_expands_dimension_for_default_row_readers(workspace_db):
    from io import BytesIO
    from zipfile import ZipFile
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font

    workbook = Workbook()
    workbook.active["A1"] = "base"
    workbook.active["A1"].font = Font(bold=True)
    output = BytesIO()
    workbook.save(output)
    base = output.getvalue()
    workbook.active["Z100"] = "AI added cell"
    output = BytesIO()
    workbook.save(output)
    proposal = output.getvalue()
    workbook = load_workbook(BytesIO(base))
    workbook.active["A1"] = "human edit"
    output = BytesIO()
    workbook.save(output)
    human = output.getvalue()

    db, workspace = workspace_db
    store = FakeStore()
    service = ServiceUnderTest(db, workspace, store)
    created = await service.upload_file("user-1", None, "report.xlsx", base)
    entry_id, base_id = created.entry.entry_id, created.entry.current_version_id
    await service.write_content("user-1", entry_id, human, expected_revision=1)
    source = "cron/runs/run-1/report.xlsx"
    store.external[("/home/user", source)] = (proposal, f"v1:{len(proposal)}:99")

    result = await service.publish_sandbox_file(
        "user-1", source_path=f"/home/user/{source}", destination_parent_id=None,
        destination_name="report.xlsx", conflict_policy="overwrite",
        expected_destination_revision=1, base_version_id=base_id,
        actor="cron", idempotency_key="publish-expanded-sheet",
    )

    assert result.status == "APPLIED"
    merged = store.files["report.xlsx"]
    parsed = load_workbook(BytesIO(merged), read_only=True)
    try:
        rows = list(parsed.active.values)
        assert rows[0][0] == "human edit"
        assert len(rows) == 100 and len(rows[-1]) == 26
        assert rows[99][25] == "AI added cell"
        assert parsed.active["A1"].font.bold is True
    finally:
        parsed.close()
    with ZipFile(BytesIO(human)) as original_zip, ZipFile(BytesIO(merged)) as merged_zip:
        assert original_zip.namelist() == merged_zip.namelist()
        for name in original_zip.namelist():
            if name != "xl/worksheets/sheet1.xml":
                assert original_zip.read(name) == merged_zip.read(name)


@pytest.mark.asyncio
async def test_agent_publish_overlap_keeps_human_version_and_silently_resolves(workspace_db):
    class ActivePathReadFailsStore(FakeStore):
        async def read_bytes(self, path, *, allow_system=False):
            if not allow_system:
                raise WorkspaceError(503, "SANDBOX_READ_FAILED", "active path stream failed")
            return await super().read_bytes(path, allow_system=allow_system)

    db, workspace = workspace_db
    store = ActivePathReadFailsStore()
    service = ServiceUnderTest(db, workspace, store)
    created = await service.upload_file("user-1", None, "shared.md", b"base")
    base_version_id = created.entry.current_version_id
    await service.write_content(
        "user-1",
        created.entry.entry_id,
        b"human edit",
        expected_revision=created.entry.revision,
    )
    # Simulate a stale materialized projection. The immutable DB head remains
    # the human edit and must repair the working file before the merge.
    store.files["shared.md"] = b"base"
    proposal = b"agent edit"
    proposal_revision = f"v1:{len(proposal)}:99"
    store.external[("/home/user", "cron/runs/run-1/candidate.md")] = (
        proposal,
        proposal_revision,
    )
    db.expire_on_commit = True

    result = await service.publish_sandbox_file(
        "user-1",
        source_path="/home/user/cron/runs/run-1/candidate.md",
        destination_parent_id=None,
        destination_name="shared.md",
        conflict_policy="overwrite",
        expected_destination_revision=1,
        base_version_id=base_version_id,
        actor="cron",
        context={"cron_run_id": "run-1"},
        idempotency_key="workspace-tool:cron:round:call",
    )

    assert result.status == "APPLIED"
    assert store.files["shared.md"] == b"human edit"
    row = db.get(WorkspaceChangeSet, result.change_set_id)
    details = json.loads(row.details_json)
    assert row.status == "applied"
    assert details["auto_merge"]["outcome"] == "current_wins"
    assert details["auto_merge"]["preserved_conflicts"] == 1
    assert store.files[details["proposal_path"]] == proposal
    current = await service.get_entry("user-1", created.entry.entry_id)
    assert row.applied_version_id == current.current_version_id
    head = db.get(WorkspaceFileVersion, current.current_version_id)
    assert store.files[head.content_path] == b"human edit"
    assert hashlib.sha256(store.files["shared.md"]).hexdigest() == current.sha256 == head.sha256
    proposal_reference = db.query(WorkspaceContentReference).filter_by(
        reference_kind="change_set_proposal",
        reference_key=f"{row.change_set_id}:proposal",
    ).one()
    assert proposal_reference.blob_id == row.proposal_blob_id
    assert proposal_reference.retained_until is not None
    assert db.query(WorkspaceContentReference).filter_by(
        reference_kind="change_set_base",
        reference_key=f"{row.change_set_id}:base",
    ).count() == 0
    frozen_result_entry = result.entry
    db.expunge_all()
    assert frozen_result_entry.name == "shared.md"
    assert frozen_result_entry.revision == current.revision


@pytest.mark.asyncio
async def test_agent_publish_absorbs_newer_physical_human_edit_before_merge(workspace_db):
    db, workspace = workspace_db
    store = FakeStore()
    service = ServiceUnderTest(db, workspace, store)
    base = b"# Report\n\nHuman: old\n\nAgent: old\n"
    human = b"# Report\n\nHuman: saved\n\nAgent: old\n"
    proposal = b"# Report\n\nHuman: old\n\nAgent: generated\n"
    created = await service.upload_file("user-1", None, "shared.md", base)
    base_version_id = created.entry.current_version_id
    store.files["shared.md"] = human
    proposal_revision = f"v1:{len(proposal)}:99"
    store.external[("/home/user", "sessions/session-1/candidate.md")] = (
        proposal,
        proposal_revision,
    )

    result = await service.publish_sandbox_file(
        "user-1",
        source_path="/home/user/sessions/session-1/candidate.md",
        destination_parent_id=None,
        destination_name="shared.md",
        conflict_policy="overwrite",
        expected_destination_revision=created.entry.revision,
        base_version_id=base_version_id,
        actor="chat",
        context={"session_id": "session-1"},
        idempotency_key="absorb-human-before-agent",
    )

    expected = b"# Report\n\nHuman: saved\n\nAgent: generated\n"
    assert result.status == "APPLIED"
    assert store.files["shared.md"] == expected
    current = await service.get_entry("user-1", created.entry.entry_id)
    head = db.get(WorkspaceFileVersion, current.current_version_id)
    absorbed = db.query(WorkspaceMutation).filter_by(operation="absorb_physical_head").one()
    assert absorbed.actor == "web"
    assert hashlib.sha256(expected).hexdigest() == current.sha256 == head.sha256
    assert store.files[head.content_path] == expected




@pytest.mark.asyncio
async def test_change_set_internal_head_read_failure_ends_failed_with_proposal_retained(workspace_db):
    class InternalHeadReadFailsStore(FakeStore):
        failed_path: str | None = None

        async def read_bytes(self, path, *, allow_system=False):
            if path == self.failed_path:
                raise WorkspaceError(503, "SANDBOX_READ_FAILED", "internal stream failed")
            return await super().read_bytes(path, allow_system=allow_system)

    db, workspace = workspace_db
    store = InternalHeadReadFailsStore()
    service = ServiceUnderTest(db, workspace, store)
    created = await service.upload_file("user-1", None, "shared.md", b"base")
    base_version_id = created.entry.current_version_id
    await service.write_content("user-1", created.entry.entry_id, b"human edit", expected_revision=1)
    head = db.get(WorkspaceFileVersion, created.entry.current_version_id)
    store.failed_path = head.content_path
    proposal = b"agent edit"
    proposal_revision = f"v1:{len(proposal)}:17"
    store.external[("/home/user", "sessions/session-1/candidate.md")] = (
        proposal,
        proposal_revision,
    )

    result = await service.publish_sandbox_file(
        "user-1",
        source_path="/home/user/sessions/session-1/candidate.md",
        destination_parent_id=None,
        destination_name="shared.md",
        conflict_policy="overwrite",
        expected_destination_revision=created.entry.revision,
        base_version_id=base_version_id,
        actor="chat",
        context={"session_id": "session-1"},
        idempotency_key="failed-head-read",
    )

    row = db.get(WorkspaceChangeSet, result.change_set_id)
    assert result.status == "FAILED"
    assert row.status == "failed"
    assert row.error_code == "SANDBOX_READ_FAILED"
    assert store.files["shared.md"] == b"human edit"
    assert store.files[json.loads(row.details_json)["proposal_path"]] == proposal
    assert await service.reconcile_change_sets("user-1") == 0


@pytest.mark.asyncio
async def test_cancelled_change_set_apply_is_converged_by_maintenance(workspace_db):
    db, workspace = workspace_db
    store = FakeStore()
    service = ServiceUnderTest(db, workspace, store)
    proposal = b"agent result"
    revision = f"v1:{len(proposal)}:7"
    store.external[("/home/user", "sessions/session-1/result.md")] = (proposal, revision)
    original_apply = service.apply_change_set
    service.apply_change_set = AsyncMock(side_effect=asyncio.CancelledError())

    with pytest.raises(asyncio.CancelledError):
        await service.publish_sandbox_file(
            "user-1",
            source_path="/home/user/sessions/session-1/result.md",
            destination_parent_id=None,
            destination_name="result.md",
            conflict_policy="fail",
            actor="chat",
            context={"session_id": "session-1"},
            idempotency_key="cancel-after-proposed",
        )

    row = db.query(WorkspaceChangeSet).filter_by(idempotency_key="cancel-after-proposed").one()
    assert row.status == "proposed"
    assert db.get(WorkspaceContentObject, row.proposal_blob_id).state == "materialized"
    service.apply_change_set = original_apply

    assert await service.reconcile_change_sets("user-1") == 1
    db.refresh(row)
    assert row.status == "applied"
    assert store.files["result.md"] == proposal

    # Publication consumes the protected object without creating another
    # platform copy outside Workspace's deletion/GC boundary.
    assert set(store.external) == {("/home/user", "sessions/session-1/result.md")}
    entry_id = row.entry_id
    entry = await service.get_entry("user-1", entry_id)
    await service.delete_entry("user-1", entry_id, expected_revision=entry.revision)
    assert not store.files
    assert store.external[("/home/user", "sessions/session-1/result.md")][0] == proposal


def test_reject_cannot_cross_applying_owner_fence(workspace_db):
    db, workspace = workspace_db
    service = ServiceUnderTest(db, workspace, FakeStore())
    change_set_id = "owned-applying-change-set"
    db.add(WorkspaceChangeSet(
        change_set_id=change_set_id,
        user_id="user-1",
        operation="publish_sandbox_file",
        status="proposed",
        actor="chat",
        idempotency_key="owned-applying",
        details_json="{}",
    ))
    db.commit()
    first_lease = service._acquire_claims(
        user_id="user-1",
        operation="apply_change_set",
        specs=(WorkspaceClaimSpec("path", f"change-set:{change_set_id}"),),
    )[0]
    row = db.get(WorkspaceChangeSet, change_set_id)
    row.status = "applying"
    row.details_json = json.dumps({
        "apply_owner": {
            "owner_token": first_lease.owner_token,
            "generation": first_lease.generation,
        },
        "apply_started_at": now_naive().isoformat(),
    }, separators=(",", ":"))
    db.commit()

    with pytest.raises(WorkspaceError) as rejected:
        service.reject_change_set("user-1", change_set_id)
    assert rejected.value.code == "CHANGE_SET_NOT_READY"
    assert db.get(WorkspaceChangeSet, change_set_id).status == "applying"
    db.rollback()
    row = db.get(WorkspaceChangeSet, change_set_id)
    row.status = "conflict"
    db.commit()
    with pytest.raises(WorkspaceError) as deferred_reject:
        service.reject_change_set("user-1", change_set_id)
    assert deferred_reject.value.code == "CHANGE_SET_NOT_READY"
    db.rollback()
    row = db.get(WorkspaceChangeSet, change_set_id)
    row.status = "applying"
    db.commit()

    service._release_unattached_claims((first_lease,))
    second_lease = service._acquire_claims(
        user_id="user-1",
        operation="apply_change_set",
        specs=(WorkspaceClaimSpec("path", f"change-set:{change_set_id}"),),
    )[0]
    row = db.get(WorkspaceChangeSet, change_set_id)
    row.details_json = json.dumps({
        "apply_owner": {
            "owner_token": second_lease.owner_token,
            "generation": second_lease.generation,
        },
        "apply_started_at": now_naive().isoformat(),
    }, separators=(",", ":"))
    db.commit()
    try:
        with pytest.raises(WorkspaceError) as fenced:
            service._finish_change_set_applied(
                change_set_id,
                apply_lease=first_lease,
                entry_id=None,
                applied_version_id=None,
            )
        assert fenced.value.code == "CHANGE_SET_APPLY_FENCED"
        assert db.get(WorkspaceChangeSet, change_set_id).status == "applying"
    finally:
        service._release_unattached_claims((second_lease,))


@pytest.mark.asyncio
async def test_cancelled_change_set_preparation_keeps_journal_and_migrates_proposal(workspace_db):
    class CancelAfterObjectStore(FakeStore):
        cancel_once = True

        async def ensure_content_object(self, **kwargs):
            result = await super().ensure_content_object(**kwargs)
            if self.cancel_once and "change-set-" in kwargs["source_relative_path"]:
                self.cancel_once = False
                raise asyncio.CancelledError()
            return result

    db, workspace = workspace_db
    store = CancelAfterObjectStore()
    service = ServiceUnderTest(db, workspace, store)
    proposal = b"agent result"
    revision = f"v1:{len(proposal)}:11"
    store.external[("/home/user", "sessions/session-1/result.md")] = (proposal, revision)

    with pytest.raises(asyncio.CancelledError):
        await service.publish_sandbox_file(
            "user-1",
            source_path="/home/user/sessions/session-1/result.md",
            destination_parent_id=None,
            destination_name="result.md",
            conflict_policy="fail",
            actor="chat",
            context={"session_id": "session-1"},
            idempotency_key="cancel-during-preparing",
        )

    row = db.query(WorkspaceChangeSet).filter_by(idempotency_key="cancel-during-preparing").one()
    details = json.loads(row.details_json)
    assert row.status == "preparing"
    assert row.proposal_blob_id is None
    proposal_temp_path = details["proposal_temp_path"]
    assert proposal_temp_path in store.files
    # Match the legacy journal shape where proposal_path still named the temp.
    details["proposal_path"] = proposal_temp_path
    details.pop("proposal_temp_path", None)
    details.pop("planned_proposal_blob_id", None)
    details.pop("planned_proposal_path", None)
    row.details_json = json.dumps(details, ensure_ascii=False, separators=(",", ":"))
    db.commit()

    assert await service.reconcile_change_sets("user-1") == 1
    db.refresh(row)
    assert row.status == "applied"
    assert row.proposal_blob_id is not None
    assert proposal_temp_path not in store.files
    assert db.query(WorkspaceContentReference).filter_by(
        reference_kind="change_set_proposal",
        reference_key=f"{row.change_set_id}:proposal",
    ).count() == 1
    assert store.files["result.md"] == proposal


@pytest.mark.asyncio
async def test_lost_preparing_change_set_reconciles_to_failed_terminal(workspace_db):
    db, workspace = workspace_db
    store = FakeStore()
    service = ServiceUnderTest(db, workspace, store)
    expected = b"lost proposal"
    change_set_id = "lost-preparing-change-set"
    db.add(WorkspaceChangeSet(
        change_set_id=change_set_id,
        user_id="user-1",
        operation="publish_sandbox_file",
        status="preparing",
        actor="chat",
        idempotency_key="lost-preparing",
        details_json=json.dumps({
            "source_sha256": hashlib.sha256(expected).hexdigest(),
            "source_size_bytes": len(expected),
            "proposal_temp_path": ".opencapybox/tmp/change-set-lost.proposal",
            "destination_name": "lost.md",
            "destination_path": "lost.md",
        }, separators=(",", ":")),
    ))
    db.commit()

    assert await service.reconcile_change_sets("user-1") == 0
    row = db.get(WorkspaceChangeSet, change_set_id)
    assert row.status == "failed"
    assert row.error_code == "CHANGE_SET_PREPARATION_LOST"






@pytest.mark.asyncio
async def test_directory_depth_limit_blocks_third_level_create_and_move(workspace_db):
    db, workspace = workspace_db
    store = FakeStore()
    service = ServiceUnderTest(db, workspace, store)
    first = await service.create_directory("user-1", None, "一级")
    second = await service.create_directory("user-1", first.entry.entry_id, "二级")

    with pytest.raises(WorkspaceError) as create_error:
        await service.create_directory("user-1", second.entry.entry_id, "三级")
    assert create_error.value.code == "DIRECTORY_DEPTH_LIMIT"

    leaf = await service.create_directory("user-1", None, "待移动")
    with pytest.raises(WorkspaceError) as leaf_move_error:
        await service.move_entry(
            "user-1",
            leaf.entry.entry_id,
            parent_id=second.entry.entry_id,
            expected_revision=leaf.entry.revision,
        )
    assert leaf_move_error.value.code == "DIRECTORY_DEPTH_LIMIT"

    destination = await service.create_directory("user-1", None, "目标")
    with pytest.raises(WorkspaceError) as subtree_move_error:
        await service.move_entry(
            "user-1",
            first.entry.entry_id,
            parent_id=destination.entry.entry_id,
            expected_revision=first.entry.revision,
        )
    assert subtree_move_error.value.code == "DIRECTORY_DEPTH_LIMIT"


class AsyncChunkReader:
    def __init__(self, chunks):
        self.chunks = list(chunks)

    async def read(self, _size):
        return self.chunks.pop(0) if self.chunks else b""


@pytest.mark.asyncio
async def test_upload_stream_keeps_chunks_bounded_and_updates_quota(workspace_db):
    db, workspace = workspace_db
    store = FakeStore()

    class FinalizationClaimService(ServiceUnderTest):
        upload_claim_states = None

        def _record_mutation(self, **kwargs):
            if kwargs.get("operation") == "upload_file":
                self.upload_claim_states = [
                    state
                    for (state,) in db.query(WorkspaceClaim.state).all()
                ]
            return super()._record_mutation(**kwargs)

    service = FinalizationClaimService(db, workspace, store)
    reader = AsyncChunkReader([b"abc", b"def", b"ghi"])

    result = await service.upload_file_stream(
        "user-1",
        None,
        "data.bin",
        reader,
        declared_size=9,
    )

    assert store.stream_read_calls == 4
    assert store.files["data.bin"] == b"abcdefghi"
    assert result.entry.size_bytes == 9
    assert workspace.used_bytes == 9
    assert service.upload_claim_states == ["active"]


@pytest.mark.asyncio
async def test_session_import_no_change_conflict_and_overwrite(workspace_db):
    db, workspace = workspace_db
    store = FakeStore()
    service = ServiceUnderTest(db, workspace, store)
    db.add(Session(id="session-1", user_id="user-1", status="active"))
    db.commit()
    source = b"version-one"
    revision = f"v1:{len(source)}:100"
    store.external[("/home/user/sessions/session-1", "report.pdf")] = (source, revision)

    created = await service.import_session_file(
        "user-1",
        session_id="session-1",
        source_path="report.pdf",
        source_revision=revision,
        destination_parent_id=None,
        destination_name="report.pdf",
        idempotency_key="import-1",
    )
    assert created.status == "CREATED"

    no_change = await service.import_session_file(
        "user-1",
        session_id="session-1",
        source_path="report.pdf",
        source_revision=revision,
        destination_parent_id=None,
        destination_name="report.pdf",
        idempotency_key="import-2",
    )
    assert no_change.status == "NO_CHANGE"

    source2 = b"version-two"
    revision2 = f"v1:{len(source2)}:200"
    store.external[("/home/user/sessions/session-1", "report.pdf")] = (source2, revision2)
    with pytest.raises(WorkspaceError) as conflict:
        await service.import_session_file(
            "user-1",
            session_id="session-1",
            source_path="report.pdf",
            source_revision=revision2,
            destination_parent_id=None,
            destination_name="report.pdf",
            idempotency_key="import-3",
        )
    assert conflict.value.code == "NAME_CONFLICT"

    overwritten = await service.import_session_file(
        "user-1",
        session_id="session-1",
        source_path="report.pdf",
        source_revision=revision2,
        destination_parent_id=None,
        destination_name="report.pdf",
        conflict_policy="overwrite",
        expected_destination_revision=created.entry.revision,
        idempotency_key="import-4",
    )
    assert overwritten.status == "UPDATED"
    assert store.files["report.pdf"] == source2


@pytest.mark.asyncio
async def test_stage_entry_freezes_current_head_or_explicit_version(workspace_db):
    db, workspace = workspace_db
    store = FakeStore()
    service = ServiceUnderTest(db, workspace, store)
    created = await service.upload_file("user-1", None, "facts.csv", b"a,b\n1,2")
    first_version_id = created.entry.current_version_id
    updated = await service.write_content(
        "user-1",
        created.entry.entry_id,
        b"a,b\n3,4",
        created.entry.revision,
    )

    staged_current = await service.stage_entry(
        "user-1",
        created.entry.entry_id,
        expected_revision=None,
        destination_root="/home/user/sessions/session-1",
    )
    staged_first = await service.stage_entry(
        "user-1",
        created.entry.entry_id,
        expected_revision=None,
        version_id=first_version_id,
        destination_root="/home/user/sessions/session-1",
    )

    assert isinstance(staged_current, WorkspaceStageResult)
    assert staged_current.source_revision == updated.entry.revision == 2
    assert staged_current.version_id == updated.entry.current_version_id
    assert staged_current.version_sequence == 2
    assert staged_first.source_revision == 2
    assert staged_first.version_id == first_version_id
    assert staged_first.version_sequence == 1
    assert store.external[("/home/user/sessions/session-1", staged_current.destination_relative_path)][0] == b"a,b\n3,4"
    assert store.external[("/home/user/sessions/session-1", staged_first.destination_relative_path)][0] == b"a,b\n1,2"


@pytest.mark.asyncio
async def test_stage_directory_preserves_one_directory_snapshot_with_empty_and_nested_entries(workspace_db):
    db, workspace = workspace_db
    store = FakeStore()
    service = ServiceUnderTest(db, workspace, store)
    folder = await service.create_directory(
        "user-1",
        None,
        "research",
        idempotency_key="mkdir-research",
    )
    await service.create_directory(
        "user-1",
        folder.entry.entry_id,
        "empty",
        idempotency_key="mkdir-empty",
    )
    nested_file = await service.upload_file(
        "user-1",
        folder.entry.entry_id,
        "facts.csv",
        b"a,b\n1,2",
    )

    staged = await service.stage_entry(
        "user-1",
        folder.entry.entry_id,
        expected_revision=folder.entry.revision,
        destination_root="/home/user/sessions/session-1",
    )

    assert staged.entry.kind == "directory"
    assert staged.source_revision == folder.entry.revision == 1
    assert staged.tree_revision == staged.entry.tree_revision
    assert staged.tree_revision > staged.source_revision
    assert staged.destination_relative_path.endswith("/research")
    assert staged.size_bytes == len(b"a,b\n1,2")
    assert staged.sha256 is not None and len(staged.sha256) == 64
    assert ("/home/user/sessions/session-1", staged.destination_relative_path) in store.external_directories
    assert (
        "/home/user/sessions/session-1",
        f"{staged.destination_relative_path}/empty",
    ) in store.external_directories
    assert store.external[(
        "/home/user/sessions/session-1",
        f"{staged.destination_relative_path}/facts.csv",
    )][0] == b"a,b\n1,2"
    assert nested_file.entry.entry_id


@pytest.mark.asyncio
async def test_stage_directory_failure_cleans_incoming_without_final_name(workspace_db):
    class FailingStageStore(FakeStore):
        async def copy_to_external_atomic(self, **kwargs):
            if kwargs["source_relative_path"].endswith("facts.csv"):
                raise WorkspaceError(503, "IO_ERROR", "copy failed")
            return await super().copy_to_external_atomic(**kwargs)

    db, workspace = workspace_db
    store = FailingStageStore()
    service = ServiceUnderTest(db, workspace, store)
    folder = await service.create_directory("user-1", None, "research")
    await service.upload_file(
        "user-1",
        folder.entry.entry_id,
        "facts.csv",
        b"a,b\n1,2",
    )

    with pytest.raises(WorkspaceError) as failed:
        await service.stage_entry(
            "user-1",
            folder.entry.entry_id,
            expected_revision=folder.entry.revision,
            destination_root="/home/user/sessions/session-1",
        )

    assert failed.value.code == "IO_ERROR"
    assert store.external == {}
    assert store.external_directories == set()


def test_non_empty_workspace_blocks_profile_switch(workspace_db):
    db, workspace = workspace_db
    workspace.entry_count = 1
    db.commit()
    with pytest.raises(WorkspaceError) as blocked:
        ensure_workspace_profile_switch_allowed(
            db,
            user_id="user-1",
            desired_profile_id="profile-2",
        )
    assert blocked.value.code == "WORKSPACE_PROFILE_SWITCH_BLOCKED"


def test_non_empty_workspace_also_blocks_same_profile_force_rebuild(workspace_db):
    db, workspace = workspace_db
    workspace.entry_count = 1
    db.commit()
    with pytest.raises(WorkspaceError) as blocked:
        ensure_workspace_profile_switch_allowed(
            db,
            user_id="user-1",
            desired_profile_id="profile-1",
        )
    assert blocked.value.code == "WORKSPACE_PROFILE_SWITCH_BLOCKED"


@pytest.mark.asyncio
async def test_prepared_write_reconciles_replace_after_database_failure(workspace_db):
    db, workspace = workspace_db
    store = FakeStore()
    normal = ServiceUnderTest(db, workspace, store)
    created = await normal.create_file("user-1", None, "note.md", file_type="markdown")
    crash = CrashAfterFilesystemService(db, workspace, store)

    with pytest.raises(RuntimeError, match="database outage"):
        await crash.write_content(
            "user-1",
            created.entry.entry_id,
            b"durable-new-content",
            created.entry.revision,
        )

    db.expire_all()
    before = db.query(WorkspaceEntry).filter(WorkspaceEntry.entry_id == created.entry.entry_id).one()
    assert before.revision == 1
    assert before.sha256 != hashlib.sha256(b"durable-new-content").hexdigest()
    prepared = db.query(WorkspaceMutation).filter(WorkspaceMutation.state == "prepared").one()
    assert prepared.after_sha256 == hashlib.sha256(b"durable-new-content").hexdigest()

    expire_prepared_mutation_ownership(db)
    assert await normal.reconcile_prepared_mutations("user-1", force=True) == 1
    db.expire_all()
    after = db.query(WorkspaceEntry).filter(WorkspaceEntry.entry_id == created.entry.entry_id).one()
    assert after.revision == 2
    assert after.sha256 == hashlib.sha256(b"durable-new-content").hexdigest()
    assert db.query(WorkspaceMutation).filter(WorkspaceMutation.mutation_id == prepared.mutation_id).one().state == "completed"


@pytest.mark.asyncio
async def test_snapshot_failure_after_install_stays_prepared_and_reconciles(workspace_db):
    db, workspace = workspace_db

    class SnapshotFailureStore(FakeStore):
        fail_next_snapshot = False

        async def copy_version_snapshot(self, **kwargs):
            if self.fail_next_snapshot:
                self.fail_next_snapshot = False
                raise WorkspaceError(404, "NOT_FOUND", "version directory unavailable")
            return await super().copy_version_snapshot(**kwargs)

    store = SnapshotFailureStore()
    service = ServiceUnderTest(db, workspace, store)
    created = await service.upload_file("user-1", None, "snapshot.md", b"old")
    entry_id = created.entry.entry_id
    store.fail_next_snapshot = True

    with pytest.raises(WorkspaceError) as failed_snapshot:
        await service.write_content(
            "user-1",
            entry_id,
            b"new",
            expected_revision=created.entry.revision,
        )
    assert failed_snapshot.value.code == "NOT_FOUND"
    assert store.files["snapshot.md"] == b"new"
    prepared = db.query(WorkspaceMutation).filter(
        WorkspaceMutation.entry_id == entry_id,
        WorkspaceMutation.state == "prepared",
    ).one()

    expire_prepared_mutation_ownership(db)
    assert await service.reconcile_prepared_mutations("user-1", force=True) == 1
    db.expire_all()
    entry = db.get(WorkspaceEntry, entry_id)
    assert entry.revision == 2
    assert entry.sha256 == hashlib.sha256(b"new").hexdigest()
    assert db.get(WorkspaceMutation, prepared.mutation_id).state == "completed"
    assert db.query(WorkspaceFileVersion).filter_by(entry_id=entry_id).count() == 2


@pytest.mark.asyncio
async def test_prepared_move_reconciles_rename_after_database_failure(workspace_db):
    db, workspace = workspace_db
    store = FakeStore()
    normal = ServiceUnderTest(db, workspace, store)
    created = await normal.upload_file("user-1", None, "old.txt", b"payload")
    crash = CrashAfterFilesystemService(db, workspace, store)

    with pytest.raises(RuntimeError, match="database outage"):
        await crash.move_entry(
            "user-1",
            created.entry.entry_id,
            parent_id=None,
            name="new.txt",
            expected_revision=created.entry.revision,
        )

    db.expire_all()
    assert db.query(WorkspaceEntry).filter(WorkspaceEntry.entry_id == created.entry.entry_id).one().relative_path == "old.txt"
    assert "new.txt" in store.files and "old.txt" not in store.files
    expire_prepared_mutation_ownership(db)
    assert await normal.reconcile_prepared_mutations("user-1", force=True) == 1
    db.expire_all()
    assert db.query(WorkspaceEntry).filter(WorkspaceEntry.entry_id == created.entry.entry_id).one().relative_path == "new.txt"


@pytest.mark.asyncio
async def test_batch_idempotency_rejects_a_different_request_fingerprint(workspace_db):
    db, workspace = workspace_db
    service = ServiceUnderTest(db, workspace, FakeStore())
    first = await service.upload_file("user-1", None, "first.md", b"first")
    second = await service.upload_file("user-1", None, "second.md", b"second")
    await service.delete_entries_batch(
        "user-1",
        ((first.entry.entry_id, first.entry.revision),),
        idempotency_key="batch-reused-key",
    )

    with pytest.raises(WorkspaceError) as reused:
        await service.delete_entries_batch(
            "user-1",
            ((second.entry.entry_id, second.entry.revision),),
            idempotency_key="batch-reused-key",
        )
    assert reused.value.code == "IDEMPOTENCY_KEY_REUSED"


@pytest.mark.asyncio
async def test_workspace_refuses_non_persistent_sandbox_before_connecting():
    service = object.__new__(WorkspaceService)
    service.settings = SimpleNamespace(
        sandbox_persistent_storage_enabled=False,
    )
    with pytest.raises(WorkspaceError) as disabled:
        await service._prepare("user-1", for_update=False)
    assert disabled.value.code == "WORKSPACE_PERSISTENCE_DISABLED"






@pytest.mark.asyncio
async def test_active_run_workspace_reconnects_exact_sandbox_without_rebuild():
    WorkspaceService._sandbox_validation_deadlines.clear()
    exact = SimpleNamespace(id="sandbox-current")
    get_existing = AsyncMock(return_value=exact)
    rebuild = AsyncMock()
    sandbox_service = SimpleNamespace(
        get_cached=lambda _user_id: None,
        invalidate_cache=lambda _user_id: None,
        get_existing=get_existing,
        get_or_resume_with_persisted_id=rebuild,
    )
    service = WorkspaceService(
        SandboxLookupDB("sandbox-current"),
        sandbox_service=sandbox_service,
    )

    result = await service._sandbox_for_user(
        "user-1",
        persisted_id="sandbox-current",
        binding_loaded=True,
        active_run=True,
    )

    assert result is exact
    get_existing.assert_awaited_once_with("user-1", "sandbox-current")
    rebuild.assert_not_awaited()


@pytest.mark.asyncio
async def test_agent_workspace_uses_run_bound_sandbox_without_global_lifecycle_lookup():
    bound = SimpleNamespace(id="sandbox-run-bound")
    sandbox_service = SimpleNamespace(
        get_cached=lambda _user_id: None,
        get_or_resume_with_persisted_id=AsyncMock(),
    )
    service = WorkspaceService(
        SandboxLookupDB("sandbox-other"),
        sandbox_service=sandbox_service,
        sandbox=bound,
    )

    result = await service._sandbox_for_user("user-1")

    assert result is bound
    sandbox_service.get_or_resume_with_persisted_id.assert_not_awaited()


@pytest.mark.asyncio
async def test_nonempty_workspace_profile_version_mismatch_fails_before_sandbox_rebuild(workspace_db):
    db, workspace = workspace_db
    workspace.entry_count = 1
    db.commit()
    service = ProfileFenceService(
        db,
        SimpleNamespace(
            profile_id="profile-1",
            profile_version=2,
            mount_path="/home/user",
        ),
    )

    with pytest.raises(WorkspaceError) as mismatch:
        await service._prepare("user-1", for_update=False)

    assert mismatch.value.code == "WORKSPACE_PROFILE_MISMATCH"
    assert service.sandbox_requested is False


@pytest.mark.asyncio
async def test_active_prepared_mutation_does_not_block_unrelated_workspace_prepare(
    workspace_db,
    monkeypatch,
):
    db, _workspace = workspace_db
    db.expire_on_commit = True
    db.add(WorkspaceMutation(
        mutation_id="prepared-other-entry",
        user_id="user-1",
        entry_id="entry-other",
        actor="web",
        operation="write_content",
        state="prepared",
        result_status="UPDATED",
        details_json='{"journal":{"bytes_delta":0,"entries_delta":0}}',
        lease_expires_at=now_naive() + timedelta(seconds=120),
    ))
    db.commit()
    service = WorkspaceService(db, sandbox_service=FakeSandboxService())
    service._runtime = lambda _user_id: SimpleNamespace(
        profile_id="profile-1",
        profile_version=1,
        mount_path="/home/user",
    )
    service._sandbox_for_user = AsyncMock(return_value=SimpleNamespace())

    async def ensure_root_noop(_store):
        return None

    monkeypatch.setattr(WorkspaceStore, "ensure_root", ensure_root_noop)

    workspace, store = await service._prepare("user-1", for_update=True)

    assert workspace.user_id == "user-1"
    assert isinstance(store, WorkspaceStore)


@pytest.mark.asyncio
async def test_active_prepared_mutations_reserve_positive_quota(workspace_db):
    db, workspace = workspace_db
    workspace.quota_bytes = 10
    db.add(WorkspaceMutation(
        mutation_id="prepared-capacity-reservation",
        user_id="user-1",
        entry_id="entry-reserved",
        actor="web",
        operation="upload_file",
        state="prepared",
        result_status="CREATED",
        details_json='{"journal":{"bytes_delta":8,"entries_delta":0}}',
        lease_expires_at=now_naive() + timedelta(seconds=120),
    ))
    db.commit()
    service = ServiceUnderTest(db, workspace, FakeStore())

    with pytest.raises(WorkspaceError) as quota:
        await service._create_or_upload_file(
            "user-1",
            None,
            "parallel.bin",
            b"1234",
            actor="web",
            idempotency_key=None,
            context=None,
            operation="create_file",
        )

    assert quota.value.code == "QUOTA_EXCEEDED"


def test_prepared_finalizers_merge_concurrent_workspace_counter_deltas(tmp_path):
    engine = create_engine(f"sqlite:///{tmp_path / 'workspace-concurrency.sqlite'}")
    UserWorkspace.__table__.create(engine)
    WorkspaceEntry.__table__.create(engine)
    WorkspaceContentReference.__table__.create(engine)
    WorkspaceMutation.__table__.create(engine)
    factory = sessionmaker(bind=engine, expire_on_commit=False)
    seed = factory()
    seed.add(UserWorkspace(
        user_id="user-concurrent",
        root_path="/home/user/workdir",
        active_profile_id="profile-1",
        active_profile_version=1,
        quota_bytes=1024,
    ))
    seed.commit()
    seed.close()
    db1 = factory()
    db2 = factory()
    try:
        ws1 = db1.query(UserWorkspace).filter_by(user_id="user-concurrent").one()
        service1 = ServiceUnderTest(db1, ws1, FakeStore())
        entry1 = WorkspaceEntry(
            entry_id="entry-1",
            user_id="user-concurrent",
            parent_id=None,
            parent_key="",
            name="one.md",
            kind="file",
            relative_path="one.md",
            size_bytes=1,
            sha256="a" * 64,
            revision=1,
            status="active",
        )
        prepared1 = service1._begin_prepared_mutation(
            workspace=ws1,
            entry_id=entry1.entry_id,
            actor="web",
            operation="create_file",
            result_status="CREATED",
            idempotency_key=None,
            context=None,
            before_revision=None,
            before_sha256=None,
            after_revision=1,
            after_sha256=entry1.sha256,
            journal={"bytes_delta": 1, "entries_delta": 1},
        )

        ws2 = db2.query(UserWorkspace).filter_by(user_id="user-concurrent").one()
        service2 = ServiceUnderTest(db2, ws2, FakeStore())
        entry2 = WorkspaceEntry(
            entry_id="entry-2",
            user_id="user-concurrent",
            parent_id=None,
            parent_key="",
            name="two.md",
            kind="file",
            relative_path="two.md",
            size_bytes=1,
            sha256="b" * 64,
            revision=1,
            status="active",
        )
        prepared2 = service2._begin_prepared_mutation(
            workspace=ws2,
            entry_id=entry2.entry_id,
            actor="web",
            operation="create_file",
            result_status="CREATED",
            idempotency_key=None,
            context=None,
            before_revision=None,
            before_sha256=None,
            after_revision=1,
            after_sha256=entry2.sha256,
            journal={"bytes_delta": 1, "entries_delta": 1},
        )

        # Simulate the stale per-request increments produced before each
        # finalizer. The finalizer must reload the authoritative row and merge
        # its own journal delta instead of overwriting the other completion.
        ws1.entry_count = 1
        ws1.used_bytes = 1
        ws2.entry_count = 1
        ws2.used_bytes = 1
        db1.add(entry1)
        db2.add(entry2)
        service1._record_mutation(
            workspace=ws1,
            entry=entry1,
            actor="web",
            operation="create_file",
            result_status="CREATED",
            idempotency_key=None,
            context=None,
            before_revision=None,
            before_sha256=None,
            prepared_mutation=prepared1,
        )
        service2._record_mutation(
            workspace=ws2,
            entry=entry2,
            actor="web",
            operation="create_file",
            result_status="CREATED",
            idempotency_key=None,
            context=None,
            before_revision=None,
            before_sha256=None,
            prepared_mutation=prepared2,
        )

        verify = factory()
        try:
            workspace = verify.query(UserWorkspace).filter_by(user_id="user-concurrent").one()
            assert workspace.entry_count == 2
            assert workspace.used_bytes == 2
            assert workspace.revision == 2
        finally:
            verify.close()
    finally:
        db1.close()
        db2.close()
        engine.dispose()


@pytest.mark.asyncio
async def test_same_revision_writers_only_one_reaches_store(tmp_path):
    engine = create_engine(f"sqlite:///{tmp_path / 'workspace-write-claim.sqlite'}")
    AuthUser.__table__.create(engine)
    UserWorkspace.__table__.create(engine)
    WorkspaceEntry.__table__.create(engine)
    WorkspaceContentObject.__table__.create(engine)
    WorkspaceContentReference.__table__.create(engine)
    WorkspaceFileVersion.__table__.create(engine)
    WorkspaceMutation.__table__.create(engine)
    WorkspaceClaim.__table__.create(engine)
    factory = sessionmaker(bind=engine, expire_on_commit=True)
    initial = b"initial"
    initial_sha = hashlib.sha256(initial).hexdigest()
    seed = factory()
    seed.add(AuthUser(
        user_id="user-concurrent-write",
        username="concurrent-write-user",
        auth_type="simple",
        password_hash="test",
        enabled=True,
    ))
    seed.add(UserWorkspace(
        user_id="user-concurrent-write",
        root_path="/home/user/workdir",
        active_profile_id="profile-1",
        active_profile_version=1,
        quota_bytes=1024,
        used_bytes=len(initial),
        entry_count=1,
    ))
    seed.add(WorkspaceEntry(
        entry_id="shared-entry",
        user_id="user-concurrent-write",
        parent_id=None,
        parent_key="",
        name="shared.md",
        kind="file",
        relative_path="shared.md",
        size_bytes=len(initial),
        mime_type="text/markdown",
        sha256=initial_sha,
        revision=1,
        status="active",
    ))
    seed.commit()
    seed.close()

    class BlockingStore(FakeStore):
        def __init__(self):
            super().__init__()
            self.files["shared.md"] = initial
            self.entered = asyncio.Event()
            self.release = asyncio.Event()
            self.write_calls = 0

        async def write_bytes_atomic(self, *args, **kwargs):
            self.write_calls += 1
            self.entered.set()
            await self.release.wait()
            return await super().write_bytes_atomic(*args, **kwargs)

    store = BlockingStore()
    db1 = factory()
    db2 = factory()
    try:
        service1 = ServiceUnderTest(db1, None, store)
        service2 = ServiceUnderTest(db2, None, store)
        first = asyncio.create_task(service1.write_content(
            "user-concurrent-write",
            "shared-entry",
            b"first writer",
            expected_revision=1,
        ))
        await store.entered.wait()

        with pytest.raises(WorkspaceError) as conflict:
            await service2.write_content(
                "user-concurrent-write",
                "shared-entry",
                b"second writer",
                expected_revision=1,
            )
        assert conflict.value.code == "WORKSPACE_MUTATION_IN_PROGRESS"
        assert store.write_calls == 1

        store.release.set()
        result = await first
        assert result.entry.revision == 2
        assert store.files["shared.md"] == b"first writer"

        verify = factory()
        try:
            entry = verify.get(WorkspaceEntry, "shared-entry")
            assert entry.revision == 2
            assert entry.sha256 == hashlib.sha256(b"first writer").hexdigest()
            assert verify.query(WorkspaceClaim).filter_by(state="active").count() == 0
        finally:
            verify.close()
    finally:
        store.release.set()
        db1.close()
        db2.close()
        engine.dispose()


def test_old_owner_cannot_finalize_or_fail_after_reconciler_takeover(workspace_db):
    db, workspace = workspace_db
    entry = WorkspaceEntry(
        entry_id="fenced-entry",
        user_id="user-1",
        parent_id=None,
        parent_key="",
        name="fenced.md",
        kind="file",
        relative_path="fenced.md",
        size_bytes=3,
        mime_type="text/markdown",
        sha256=hashlib.sha256(b"old").hexdigest(),
        revision=1,
        status="active",
    )
    db.add(entry)
    workspace.entry_count = 1
    workspace.used_bytes = 3
    db.commit()
    service = ServiceUnderTest(db, workspace, FakeStore())
    entry_id = entry.entry_id
    before = service._journal_projection(entry)
    after = {**before, "revision": 2, "sha256": hashlib.sha256(b"new").hexdigest()}
    db.rollback()
    prepared = service._begin_prepared_mutation(
        workspace=workspace,
        workspace_user_id="user-1",
        entry_id=entry_id,
        actor="web",
        operation="write_content",
        result_status="UPDATED",
        idempotency_key=None,
        context=None,
        before_revision=1,
        before_sha256=before["sha256"],
        after_revision=2,
        after_sha256=after["sha256"],
        journal={
            "target_path": "fenced.md",
            "old_sha256": before["sha256"],
            "new_sha256": after["sha256"],
            "bytes_delta": 0,
            "entries_delta": 0,
            "before_entry_projection": before,
            "entry_projection": after,
        },
        claim_specs=(WorkspaceClaimSpec(
            "file",
            file_scope(entry_id),
            entry_id,
        ),),
    )
    expire_prepared_mutation_ownership(db)
    takeover = WorkspaceMutationCoordinator(db).takeover_expired_mutation_claims(
        user_id="user-1",
        mutation_id=prepared.mutation_id,
    )
    replacement = takeover.current
    assert takeover.previous[0].owner_token == prepared.leases[0].owner_token
    assert replacement[0].generation == takeover.previous[0].generation + 1

    with pytest.raises(WorkspaceError) as fenced:
        service._record_mutation(
            workspace=workspace,
            entry=entry,
            actor="web",
            operation="write_content",
            result_status="UPDATED",
            idempotency_key=None,
            context=None,
            before_revision=1,
            before_sha256=before["sha256"],
            prepared_mutation=prepared,
        )
    assert fenced.value.code == "MUTATION_FENCED"
    service._fail_prepared_mutation(
        prepared,
        code="LATE_FAILURE",
        message="old owner returned late",
        recoverable=True,
    )
    mutation = db.get(WorkspaceMutation, prepared.mutation_id)
    assert mutation.state == "prepared"
    assert mutation.owner_token == replacement[0].owner_token


def test_failed_idempotent_replay_returns_frozen_failure(workspace_db):
    db, _workspace = workspace_db
    mutation = WorkspaceMutation(
        mutation_id="failed-replay-mutation",
        user_id="user-1",
        entry_id="failed-entry",
        actor="web",
        operation="write_content",
        state="failed",
        result_status="UPDATED",
        idempotency_key="failed-replay-key",
        details_json=json.dumps({
            "failure": {
                "status_code": 412,
                "code": "DESTINATION_CHANGED",
                "message": "目标文件已变化",
                "extra": {"sha256": "new-sha"},
            },
        }),
        error_code="DESTINATION_CHANGED",
        error_message="目标文件已变化",
        recoverable=True,
        completed_at=now_naive(),
    )
    db.add(mutation)
    db.commit()
    service = WorkspaceService(db, sandbox_service=SimpleNamespace())

    with pytest.raises(WorkspaceError) as replayed:
        service._idempotent_result(
            "user-1",
            "failed-replay-key",
            "write_content",
        )

    assert replayed.value.status_code == 412
    assert replayed.value.code == "DESTINATION_CHANGED"
    assert replayed.value.message == "目标文件已变化"
    assert replayed.value.extra == {
        "sha256": "new-sha",
        "mutation_id": "failed-replay-mutation",
        "mutation_state": "failed",
        "outcome": "not_applied",
    }


@pytest.mark.asyncio
async def test_workspace_metadata_listing_never_resumes_sandbox(workspace_db):
    db, _workspace = workspace_db
    service = ProfileFenceService(
        db,
        SimpleNamespace(
            profile_id="profile-1",
            profile_version=1,
            mount_path="/home/user",
        ),
    )

    page = await service.list_entries("user-1")

    assert page.items == []
    assert service.sandbox_requested is False


@pytest.mark.parametrize(
    ("target_name", "wildcard_sibling_name"),
    [
        ("fund%2026", "fundX2026"),
        ("team_1", "teamA1"),
    ],
)
@pytest.mark.asyncio
async def test_directory_mutations_escape_like_metacharacters_for_descendants(
    workspace_db,
    target_name,
    wildcard_sibling_name,
):
    db, workspace = workspace_db
    store = FakeStore()
    service = ServiceUnderTest(db, workspace, store)

    source_parent = await service.create_directory("user-1", None, "source")
    destination_parent = await service.create_directory("user-1", None, "destination")
    target = await service.create_directory(
        "user-1",
        source_parent.entry.entry_id,
        target_name,
    )
    target_child = await service.create_file(
        "user-1",
        target.entry.entry_id,
        "target-child.md",
        file_type="markdown",
    )

    source_sibling = await service.create_directory(
        "user-1",
        source_parent.entry.entry_id,
        wildcard_sibling_name,
    )
    source_sibling_child = await service.create_file(
        "user-1",
        source_sibling.entry.entry_id,
        "source-sibling.md",
        file_type="markdown",
    )
    destination_sibling = await service.create_directory(
        "user-1",
        destination_parent.entry.entry_id,
        wildcard_sibling_name,
    )
    destination_sibling_child = await service.create_file(
        "user-1",
        destination_sibling.entry.entry_id,
        "destination-sibling.md",
        file_type="markdown",
    )
    target_entry_id = target.entry.entry_id
    target_child_entry_id = target_child.entry.entry_id
    source_sibling_child_id = source_sibling_child.entry.entry_id
    destination_sibling_child_id = destination_sibling_child.entry.entry_id

    source_sibling_path = (
        f"source/{wildcard_sibling_name}/source-sibling.md"
    )
    destination_sibling_path = (
        f"destination/{wildcard_sibling_name}/destination-sibling.md"
    )

    moved = await service.move_entry(
        "user-1",
        target_entry_id,
        parent_id=destination_parent.entry.entry_id,
        expected_revision=target.entry.revision,
    )
    db.expire_all()
    untouched_source = db.query(WorkspaceEntry).filter(
        WorkspaceEntry.entry_id == source_sibling_child_id
    ).one()
    assert untouched_source.relative_path == source_sibling_path
    assert untouched_source.status == "active"
    assert untouched_source.revision == 1

    await service.delete_entry(
        "user-1", target_entry_id, expected_revision=moved.entry.revision,
    )
    db.expire_all()
    assert db.query(WorkspaceEntry).filter(
        WorkspaceEntry.entry_id.in_((target_entry_id, target_child_entry_id))
    ).count() == 0
    for sibling_id, expected_path in (
        (source_sibling_child_id, source_sibling_path),
        (destination_sibling_child_id, destination_sibling_path),
    ):
        untouched = db.query(WorkspaceEntry).filter(
            WorkspaceEntry.entry_id == sibling_id
        ).one()
        assert untouched.relative_path == expected_path
        assert untouched.status == "active"
        assert untouched.revision == 1

@pytest.mark.asyncio
async def test_direct_delete_removes_history_keeps_shared_bytes_and_replays_frozen_result(workspace_db):
    db, workspace = workspace_db
    store = FakeStore()
    service = ServiceUnderTest(db, workspace, store)
    a = await service.upload_file("user-1", None, "a.md", b"shared")
    a_id, a_version = a.entry.entry_id, a.entry.current_version_id
    b = await service.upload_file("user-1", None, "b.md", b"shared")
    service.protect_version_reference("user-1", a_version, reference_kind="round_attachment", reference_key="s:r")
    await service.write_content("user-1", a_id, b"draft-one", 1)
    await service.write_content("user-1", a_id, b"draft-two", 2)
    store.remove_content_objects = AsyncMock(wraps=store.remove_content_objects)
    store.remove_office_preview_caches = AsyncMock(
        wraps=store.remove_office_preview_caches
    )
    request = ((a_id, 3),)
    result = await service.delete_entries_batch("user-1", request, idempotency_key="delete-a")
    assert result.affected_entry_ids == (a_id,)
    assert db.get(WorkspaceEntry, a_id) is None
    assert db.get(WorkspaceFileVersion, a_version) is None
    assert db.query(WorkspaceContentReference).filter_by(version_id=a_version).count() == 0
    assert store.files["b.md"] == b"shared"
    store.remove_content_objects.assert_awaited_once()
    assert len(list(store.remove_content_objects.await_args.args[0])) == 2
    draft_cache_keys = {
        *office_preview_cache_keys(hashlib.sha256(b"draft-one").hexdigest()),
        *office_preview_cache_keys(hashlib.sha256(b"draft-two").hexdigest()),
    }
    shared_cache_keys = set(
        office_preview_cache_keys(hashlib.sha256(b"shared").hexdigest())
    )
    assert set(store.remove_office_preview_caches.await_args.args[0]) == draft_cache_keys
    assert shared_cache_keys.isdisjoint(draft_cache_keys)
    assert (await service.open_content("user-1", b.entry.entry_id)).sandbox_path
    replay = await service.delete_entries_batch("user-1", request, idempotency_key="delete-a")
    assert replay == result
    await service.delete_entry("user-1", b.entry.entry_id, expected_revision=1)
    assert set(store.remove_office_preview_caches.await_args.args[0]) == shared_cache_keys
    assert not store.files
    db.refresh(workspace)
    assert (workspace.used_bytes, workspace.entry_count, workspace.history_used_bytes) == (0, 0, 0)


@pytest.mark.asyncio
async def test_batch_object_cleanup_retries_partial_failure_without_double_accounting(workspace_db):
    db, workspace = workspace_db
    store = FakeStore()
    service = ServiceUnderTest(db, workspace, store)
    a = await service.upload_file("user-1", None, "a.md", b"first")
    b = await service.upload_file("user-1", None, "b.md", b"second")
    request = ((a.entry.entry_id, 1), (b.entry.entry_id, 1))
    remove_batch = store.remove_content_objects

    async def interrupted(paths):
        await store.remove(next(iter(paths)))
        raise WorkspaceError(503, "IO_ERROR", "interrupted")

    store.remove_content_objects = interrupted
    await service.delete_entries_batch("user-1", request, idempotency_key="delete-gc-retry")
    db.refresh(workspace)
    assert workspace.used_bytes == workspace.entry_count == 0
    assert workspace.history_used_bytes == len(b"firstsecond")
    assert db.query(WorkspaceContentObject).filter_by(state="pruning").count() == 2
    store.remove_content_objects = remove_batch
    result = await service.run_history_gc("user-1")
    assert result.objects_pruned == 2
    db.refresh(workspace)
    assert workspace.history_used_bytes == 0
    assert not store.files
    assert db.query(WorkspaceClaim).filter_by(state="active").count() == 0


@pytest.mark.asyncio
async def test_partial_direct_delete_recovers_without_deleting_a_reused_path(workspace_db):
    class PartialStore(FakeStore):
        failed = False
        async def delete_entries(self, paths, cleanup_paths):
            if not self.failed:
                self.failed = True
                await self.remove(paths[0])
                raise WorkspaceError(503, "IO_ERROR", "interrupted")
            await super().delete_entries(paths, cleanup_paths)
    db, workspace = workspace_db
    store = PartialStore()
    service = ServiceUnderTest(db, workspace, store)
    folder = await service.create_directory("user-1", None, "folder")
    child = await service.upload_file("user-1", folder.entry.entry_id, "a.md", b"a")
    second = await service.upload_file("user-1", None, "b.md", b"b")
    request = ((folder.entry.entry_id, 1), (child.entry.entry_id, 1), (second.entry.entry_id, 1))
    with pytest.raises(WorkspaceError):
        await service.delete_entries_batch("user-1", request, idempotency_key="delete-batch")
    with pytest.raises(WorkspaceError):
        await service.create_directory("user-1", None, "folder")
    expire_prepared_mutation_ownership(db)
    assert await service.reconcile_prepared_mutations("user-1", force=True) == 1
    assert db.query(WorkspaceEntry).count() == 0
    replay = await service.delete_entries_batch("user-1", request, idempotency_key="delete-batch")
    assert len(replay.roots) == 2
    assert len(replay.affected_entry_ids) == 3
    assert db.query(WorkspaceClaim).filter_by(state="active").count() == 0


@pytest.mark.asyncio
async def test_direct_delete_script_preflights_all_roots_and_rejects_system_paths():
    runner = AsyncMock(return_value=SimpleNamespace(exit_code=0, stdout='{"ok":true}'))
    store = WorkspaceStore(SimpleNamespace(commands=SimpleNamespace(run=runner)), "/home/user/workdir")
    await store.delete_entries(["folder", "a.md"], [".opencapybox/read/entry"])
    body = runner.await_args.args[0].split("<<'PY'\n", 1)[1].rsplit("\nPY", 1)[0]
    compile(body, "<direct-delete>", "exec")
    assert body.index("check_tree(fd, name)") < body.index("remove_tree(fd, name)")
    with pytest.raises(WorkspaceError):
        await store.delete_entries([".opencapybox/objects"], [])

    runner.reset_mock()
    paths = [f".opencapybox/objects/sha256/{digit * 2}/{digit * 64}/content" for digit in ("a", "b")]
    await store.remove_content_objects(paths)
    runner.assert_awaited_once()
    body = runner.await_args.args[0].split("<<'PY'\n", 1)[1].rsplit("\nPY", 1)[0]
    compile(body, "<batch-object-delete>", "exec")
    assert all(path in body for path in paths)
    assert body.index("os.stat(name") < body.index("os.unlink(name")
    with pytest.raises(WorkspaceError):
        await store.remove_content_objects([paths[0], ".opencapybox/objects"])
    runner.assert_awaited_once()

    runner.reset_mock()
    cache_keys = ["a" * 64, "b" * 64]
    await store.remove_office_preview_caches(cache_keys)
    runner.assert_awaited_once()
    cache_body = runner.await_args.args[0].split("<<'PY'\n", 1)[1].rsplit("\nPY", 1)[0]
    compile(cache_body, "<batch-preview-cache-delete>", "exec")
    assert all(cache_key in cache_body for cache_key in cache_keys)
    assert "follow_symlinks=False" in cache_body
    with pytest.raises(WorkspaceError):
        await store.remove_office_preview_caches(["not-a-cache-key"])
